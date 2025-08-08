import yaml
import io
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.utils import *
from reportlab.lib.pagesizes import *
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, white
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.pdfgen import canvas

def generate_bom_from_yaml(filename: str, yaml_text: str, inputs: dict, generate_pdf: bool) -> bytes:
    """
    Parses the YAML (which defines questions and BOM rules),
    validates the inputs, builds the BOM by looking up each product from the database,
    and writes the BOM out to an Excel workbook.
    """
    # Parse the YAML text.
    try:
        schema = yaml.safe_load(yaml_text)
    except yaml.YAMLError as e:
        raise ValueError(f"Invalid YAML: {e}")

    # Ensure the YAML structure has the required keys.
    if not isinstance(schema, dict) or 'products' not in schema:
        raise ValueError("YAML must contain a 'products' key")

    context = dict(inputs)
    
    # Validate each question and input.
    for question in schema.get("questions", []):
        name = question.get("name")
        expected_type = question.get("type")
        user_input = inputs.get(name)
        if user_input is None:
            raise ValueError(f"Missing input for {name}. Expected type: {expected_type}")

        if expected_type == "integer":
            if not isinstance(user_input, int):
                raise TypeError(f"Input for {name} must be an integer, got {type(user_input).__name__} instead")
            min_val = question.get("min", user_input)
            max_val = question.get("max", user_input)
            if int(user_input) < int(min_val):
                raise TypeError(f"Input for {name} must be >= {min_val} (got {user_input})")
            if int(user_input) > int(max_val):
                raise TypeError(f"Input for {name} must be <= {max_val} (got {user_input})")
        elif expected_type == "boolean":
            if not isinstance(user_input, bool):
                raise TypeError(f"Input for {name} must be a boolean, got {type(user_input).__name__} instead")
        elif expected_type == "enum":
            choices = question.get("choices")
            if not choices:
                raise ValueError(f"Missing choices for enum input {name}")
            if user_input not in choices:
                raise ValueError(f"Input for {name} must be one of {choices}, got {user_input}")
        else:
            raise ValueError(f"Unknown expected type: {expected_type} for input {name}")

    # Initialize the Excel workbook.
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bill of Materials for {filename}"
    
    # Set up header row corresponding to the sample format.
    headers = [
        "Manufacturer Part #",
        "Manufacturer",
        "Description",
        "Device Role",
        "Qty",
        "List Price",
        "Discount",
        "Customer Price",
        "Ext. Price",
    ]
    
    # Write header row
    ws.append(headers)
    
    # Extra empty row (just for looks)
    ws.append([""] * len(headers))
    
    # First collect all products and their quantities that are going to be in the BOM
    collective_parts = {}
    for i, rule in enumerate(schema.get("products")):
        add = rule.get("add")
        
        # If it's raw, add it directly
        raw = add.get("raw")
        if raw:
            collective_parts[f"raw_{i}"] = {"raw": raw}
        else:
            # It's a product
            cond = rule.get("condition")  # condition string is optional
            if not add or not isinstance(add, dict):
                raise ValueError(f"Rule #{i + 1} must have an 'add' dict")
            part = add.get("product")
            # Skip rules meant as examples
            if part == "example":
                continue
            
            try:
                # Evaluate the optional condition.
                if cond and not eval(cond, {}, context):
                    continue
                
                # Evaluate the quantity: if it is a string, evaluate it as a Python expression.
                quantity_val = add.get("quantity")
                if isinstance(quantity_val, str):
                    quantity = eval(quantity_val, {}, context)
                elif isinstance(quantity_val, int):
                    quantity = quantity_val
                else:
                    raise ValueError(f"Invalid quantity type for product '{part}'")
                
                # Ensure it's an int. Cause you can't have half of a router
                if type(quantity) is not int:
                    raise ValueError(f"Quantities must be integers (got {type(quantity)} for value {quantity})")
            except Exception as e:
                raise ValueError(f"Error evaluating quantity for product section #{i + 1}: {e}")
            
            # Check if we already inserted this product into a row
            if part in collective_parts:
                collective_parts[part]["quantity"] += quantity
            else:
                collective_parts[part] = {"quantity": quantity}
    
    # If there's no parts to add and this is NOT a save action (meaning someone's trying to generate a PDF)
    if generate_pdf and not collective_parts:
        # Don't bother generating a PDF, just return early
        raise ValueError("BOM is empty")
    
    # Now that we've gathered all of the collective parts, let's assign one row to each product
    subtotal_price = 0.0
    for part in collective_parts:
        # If it's a raw part
        raw = collective_parts[part].get("raw")
        if raw:
            # Ensure it won't overflow
            if len(raw) > len(headers):
                raise ValueError(f"'raw' list has too many entries (max: {len(headers)})")
            ws.append(raw)
            continue
        # Otherwise, it's a product
        # Retrieve the product object to extract metadata
        try:
            from .models import Product
            # Look up the product by its part number using the DB
            product = Product.objects.get(part=part)
        except Product.DoesNotExist:
            raise ValueError(f"Product '{part}' does not exist")
        
        # Retrieve product details
        quantity = collective_parts[part]["quantity"]
        cust_discount = float(product.discount)
        cust_price = round_currency(float(product.list_price) * (1 - cust_discount))
        ext_price = round_currency(cust_price * quantity)
        # Gets added at the bottom
        subtotal_price += ext_price
        
        # Each value in this array is one cell (left to right) in a new row
        ws.append([
            part, # Manufacturer part #
            product.manufacturer.name, # Manufacturer
            product.description, # Description
            (product.device_role and product.device_role.name) or "N/A", # Device role (optional)
            quantity,
            format_currency(float(product.list_price)),
            format_percentage(cust_discount),
            format_currency(cust_price),
            format_currency(ext_price),
        ])
    
    # Append the subtotal row
    subtotal_row = [''] * len(headers)
    subtotal_row[ws.max_column - 2] = "Subtotal:"
    subtotal_row[ws.max_column - 1] = format_currency(subtotal_price)
    ws.append(subtotal_row)
    
    # Extra empty row (just for looks, again)
    ws.append([""] * len(headers))
    
    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = ws.cell(row=1, column=col).column_letter
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # The widths of the columns are used in the PDF generation too
        adjusted_width = max_length + 2 # add some padding
        ws.column_dimensions[column].width = adjusted_width
    
    # Save as PDF
    return excel_to_pdf(wb)

# Font used in the PDF
bom_font = Font("Helvetica", size=20, color="FFFFFF")

# Heavily modified from https://github.com/rameshvoodi/excel-to-pdf-python/blob/main/main.py
# Only the width properties are used from the excel sheet, the height is managed here
def excel_to_pdf(workbook):
    pdf_output = io.BytesIO()
    
    # Only one worksheet to process
    sheet = workbook.worksheets[0]
    
    # Get number of columns and rows
    num_columns = sheet.max_column
    # Initialize column widths
    column_widths = [0] * num_columns
    
    # Calculate column widths based on the maximum length of cells in each column
    for col_index in range(1, num_columns + 1):
        max_length = 0
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_index).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        # Calculate width (x10 for the scale factor, might want to play with this) with some padding
        column_widths[col_index - 1] = ((max_length + 5) * (bom_font.size * 0.6))
    # Calculate total width
    total_width = sum(column_widths)
    # Calculate total height based on row heights
    padding_scale = 1.5
    total_height = (bom_font.size * padding_scale * sheet.max_row) * 1.33333
    
    # Set the page size to match the total width
    page_size = (total_width, total_height)
    
    c = canvas.Canvas(pdf_output, pagesize=page_size)
    
    # This is like the drawY. Init to top of page
    y = page_size[1]
    
    for index, row in enumerate(list(sheet.iter_rows())):
        if all(cell.value is None for cell in row):
            continue
        
        x = 0
        row_height = bom_font.size * padding_scale * 1.33333
        
        for column_index, cell in enumerate(row):
            cell_width = column_widths[column_index]
            # Raise it a lil
            text_y = y - row_height + (row_height - bom_font.size) / 2 + (bom_font.size / 8)
            
            # Handle cell background color on the first row
            if index == 0:
                # Set the fill color for the background
                c.setFillColor(black)
                c.rect(x, y - row_height, cell_width, row_height, fill=1)
                # Set the text color to white
                c.setFillColor(white)
            elif index == 1 or index == sheet.max_row - 1:
                grey = 0.75
                c.setFillColorRGB(grey, grey, grey)
                c.setStrokeColorRGB(grey, grey, grey)
                c.rect(x, y - row_height, cell_width, row_height, fill=1)
                # Set the text color to white
                c.setFillColor(white)
            else:
                c.setFillColor(black)
            
            # Bold + center text
            text_object = c.beginText()
            text = get_cell_value(cell)
            text_width = c.stringWidth(text, "Helvetica", bom_font.size)
            text_object.setTextOrigin(x + (cell_width - text_width) / 2, text_y)
            # Bold the first row and the 2nd to last one (the subtotal line)
            text_object.setFont(f"Helvetica{"-Bold" if index == 0 or index == sheet.max_row - 2 else ""}", bom_font.size)
            
            if index == 0:
                text_object.textLine(get_cell_value(cell))
            else:
                text_object.textLine(get_cell_value(cell))
            
            c.drawText(text_object)
            x += cell_width

        y -= row_height
    
    # Add a disclaimer
    text_object = c.beginText()
    text = "*Prices listed are estimates and may vary"
    text_object.setTextOrigin(bom_font.size / 2, ((bom_font.size * padding_scale * 1.33333) - bom_font.size) / 2 + (bom_font.size / 8))
    text_object.textLine(text)
    c.setFillColor(black)
    c.drawText(text_object)
    
    # Save and export the PDF
    c.save()
    pdf_output.seek(0)
    return pdf_output.getvalue()

def round_currency(value):
    """Rounds float to 2 decimal places (for currency calculations)"""
    return round(value, 2)

def format_currency(value):
    """Formats a number as currency"""
    if value is None:
        return "$0.00"
    return f"${value:,.2f}"

def format_percentage(value):
    """Formats a number as a percentage"""
    if value is None:
        return "0%"
    return f"{value * 100}%"

def get_cell_value(cell):
    if cell.value is None:
        return ""
    elif cell.data_type == "n":
        return str(cell.value)
    elif cell.data_type == "s":
        return cell.value
    elif cell.data_type == "b":
        return "TRUE" if cell.value else "FALSE"
    elif cell.data_type == "d":
        return cell.value.strftime("%Y-%m-%d")
    elif cell.data_type == "e":
        return f"ERROR: {cell.value}"
    elif cell.data_type == "t":
        return cell.value.strftime("%H:%M:%S")
    else:
        return str(cell.value)
